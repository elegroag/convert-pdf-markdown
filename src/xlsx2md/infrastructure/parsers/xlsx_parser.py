"""OpenPyXL-based spreadsheet parser for XLSX2MD.

Strategy
--------
Excel workbooks are heterogeneous: some contain a single big table, others are
"informes" with narrative sections followed by tabular regions. This parser
classifies each worksheet and extracts content accordingly:

* **NarrativeSheet** — alternates between headings, key/value pairs and tables.
* **TableSheet**    — the worksheet is dominated by a single table region.
* **ImageSheet**    — the worksheet is mostly a logo/image holder.
* **EmptySheet**    — the worksheet has no meaningful content.

A worksheet is classified as a table when >=80% of populated rows belong to a
single rectangular region with consistent column count. Otherwise it is treated
as a narrative sheet and broken into blocks of three kinds:

* ``HeadingBlock``   — a row with text bold/large or a single non-empty cell.
* ``KeyValueBlock``  — a row with exactly two non-empty cells ``label: value``.
* ``TableBlock``     — a rectangular run of rows with the same column count.
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from xlsx2md.domain.entities.entities import (
    EmptySheet,
    HeadingBlock,
    ImageBlock,
    ImageSheet,
    KeyValueBlock,
    NarrativeSheet,
    ParagraphBlock,
    TableBlock,
    TableSheet,
    XlsxDocument,
    XlsxMetadata,
)
from xlsx2md.domain.exceptions import CorruptedXlsxError
from xlsx2md.domain.ports.ports import IAssetExporter, ISpreadsheetParser
from xlsx2md.domain.value_objects.value_objects import ConversionConfig
from xlsx2md.infrastructure.storage.asset_exporter import FileAssetExporter


# Heuristics for block detection
TABLE_REGION_MIN_ROWS = 3
TABLE_REGION_MIN_DENSITY = 0.6
TABLE_DOMINANCE_THRESHOLD = 0.8
HEADER_KEYWORDS = (
    "item",
    "id",
    "nombre",
    "descripcion",
    "descripción",
    "fecha",
    "estado",
    "responsable",
    "actividad",
    "porcentaje",
    "cumplimiento",
    "recomendacion",
    "recomendación",
    "accion",
    "acción",
    "mejora",
    "observacion",
    "observación",
)


def _coerce_to_string(cell: object) -> str:
    """Convert any cell value into the string that should appear in Markdown."""
    value = getattr(cell, "value", None)
    if value is None or value == "":
        return ""

    hyperlink = getattr(cell, "hyperlink", None)
    if hyperlink is not None and getattr(hyperlink, "target", None):
        label = str(value) if value is not None else str(hyperlink.target)
        return f"[{label}]({hyperlink.target})"

    data_type = getattr(cell, "data_type", None)
    if data_type == "f":
        return f"`{value}`"

    if data_type == "d" or isinstance(value, datetime):
        return _format_date(value)

    if isinstance(value, bool):
        return "Sí" if value else "No"

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"

    text = str(value)
    # Collapse runs of internal whitespace (Excel often pads strings with
    # extra spaces due to legacy cell formatting) and strip ends.
    text = re.sub(r"[ \t\u00a0]+", " ", text)
    return text.strip()


def _format_date(value: datetime) -> str:
    """Render a datetime in ISO date format."""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)


class XlsxParser(ISpreadsheetParser):
    """Parse an Excel workbook into domain entities."""

    def __init__(
        self,
        asset_exporter: IAssetExporter | None = None,
        config: ConversionConfig | None = None,
    ) -> None:
        self._asset_exporter = asset_exporter
        self._config = config or ConversionConfig()

    def parse(self, path: Path, *, book_dir: Path) -> XlsxDocument:
        """Parse the workbook at ``path``."""
        if self._config.extract_images:
            assets_dir = book_dir / self._config.assets_subdir
            self._asset_exporter = FileAssetExporter(assets_dir=assets_dir, config=self._config)

        try:
            workbook_values = load_workbook(path, data_only=False, keep_vba=False)
        except Exception as exc:  # noqa: BLE001
            raise CorruptedXlsxError(f"cannot open {path}: {exc}") from exc

        try:
            metadata = self._build_metadata(workbook_values)
            sheets: list = []
            for idx, worksheet in enumerate(workbook_values.worksheets):
                sheets.append(self._parse_sheet(worksheet, idx))
            return XlsxDocument(file_path=path, sheets=sheets, metadata=metadata)
        finally:
            workbook_values.close()

    def _build_metadata(self, workbook: object) -> XlsxMetadata:
        props = getattr(workbook, "properties", None)
        return XlsxMetadata(
            title=str(getattr(props, "title", "") or ""),
            author=str(getattr(props, "creator", "") or ""),
            subject=str(getattr(props, "subject", "") or ""),
            creator=str(getattr(props, "creator", "") or ""),
            sheet_names=tuple(getattr(workbook, "sheetnames", ())),
        )

    def _parse_sheet(self, worksheet: object, index: int) -> object:
        max_row = int(getattr(worksheet, "max_row", 0) or 0)
        max_col = int(getattr(worksheet, "max_column", 0) or 0)
        raw_dimensions = str(getattr(worksheet, "dimensions", ""))

        # Prune trailing empty rows and columns so we never iterate over padding.
        max_row, max_col = self._trim_empty_bounds(worksheet, max_row, max_col)
        if max_row == 0 or max_col == 0:
            return EmptySheet(
                name=str(getattr(worksheet, "title", f"Sheet{index + 1}")),
                index=index,
                raw_dimensions=raw_dimensions,
            )

        populated_rows = self._count_populated_rows(worksheet, max_row, max_col)
        if populated_rows == 0:
            return EmptySheet(
                name=str(getattr(worksheet, "title", f"Sheet{index + 1}")),
                index=index,
                raw_dimensions=raw_dimensions,
            )

        images = self._extract_images(worksheet)

        if len(images) > 0 and populated_rows <= 3:
            return ImageSheet(
                name=str(getattr(worksheet, "title", f"Sheet{index + 1}")),
                index=index,
                images=images,
                raw_dimensions=raw_dimensions,
            )

        col_limit = self._config.max_cols or self._config.default_table_max_cols
        effective_max_col = min(max_col, col_limit)

        table_region = self._detect_single_table(
            worksheet, max_row, effective_max_col, populated_rows
        )
        if self._config.detect_blocks and table_region is not None:
            start_row, end_row, start_col, end_col = table_region
            table = self._build_table_block(worksheet, start_row, end_row, start_col, end_col)
            if not table.is_empty():
                sheet_name = str(getattr(worksheet, "title", f"Sheet{index + 1}"))
                return TableSheet(
                    name=sheet_name,
                    index=index,
                    tables=[table],
                    images=images,
                    raw_dimensions=raw_dimensions,
                )

        return self._parse_narrative_sheet(
            worksheet, index, max_row, max_col, images, raw_dimensions, effective_max_col
        )

    @staticmethod
    def _trim_empty_bounds(
        worksheet: object,
        max_row: int,
        max_col: int,
    ) -> tuple[int, int]:
        """Shrink ``max_row`` and ``max_col`` to drop trailing empty rows/cols."""
        trimmed_row = 0
        for r in range(max_row, 0, -1):
            for c in range(1, max_col + 1):
                if getattr(worksheet.cell(row=r, column=c), "value", None) not in (None, ""):
                    trimmed_row = r
                    break
            if trimmed_row:
                break

        trimmed_col = 0
        for c in range(max_col, 0, -1):
            for r in range(1, max_row + 1):
                if getattr(worksheet.cell(row=r, column=c), "value", None) not in (None, ""):
                    trimmed_col = c
                    break
            if trimmed_col:
                break

        return trimmed_row, trimmed_col

    def _parse_narrative_sheet(
        self,
        worksheet: object,
        index: int,
        max_row: int,
        max_col: int,
        images: list[ImageBlock],
        raw_dimensions: str,
        effective_max_col: int,
    ) -> NarrativeSheet:
        sheet_name = str(getattr(worksheet, "title", f"Sheet{index + 1}"))
        blocks: list = []

        rows_data = self._load_row_data(worksheet, max_row, max_col)
        runs = self._collect_row_runs(rows_data)

        for run_start, run_end, start_col, end_col in runs:
            row_count = run_end - run_start + 1
            col_count = end_col - start_col + 1
            density = self._density_for_run(rows_data, run_start, run_end, start_col, end_col)
            if (
                row_count >= TABLE_REGION_MIN_ROWS
                and col_count >= 2
                and density >= TABLE_REGION_MIN_DENSITY
            ):
                table = self._build_table_block_from_data(
                    rows_data, run_start, run_end, start_col, end_col, worksheet
                )
                if not table.is_empty():
                    blocks.append(table)
                continue

            # Group sparse rows with the same column footprint into a single table.
            grouped = self._try_group_sparse_rows(rows_data, run_start, run_end, start_col, end_col)
            if grouped is not None:
                g_start, g_end, g_start_col, g_end_col, g_pattern = grouped
                table = self._build_table_block_from_data(
                    rows_data, g_start, g_end, g_start_col, g_end_col, worksheet
                )
                if not table.is_empty():
                    blocks.append(table)
                continue

            if row_count == 1 and col_count == 2:
                label_cell = worksheet.cell(row=run_start, column=start_col)
                value_cell = worksheet.cell(row=run_start, column=start_col + 1)
                label_text = _coerce_to_string(label_cell).rstrip(":").strip()
                value_text = _coerce_to_string(value_cell)
                if label_text and value_text:
                    blocks.append(
                        KeyValueBlock(
                            label=label_text,
                            value=value_text,
                            anchor=str(getattr(label_cell, "coordinate", "")),
                        )
                    )
            elif row_count == 1 and col_count == 1:
                cell = worksheet.cell(row=run_start, column=start_col)
                text = _coerce_to_string(cell)
                if text:
                    blocks.append(
                        HeadingBlock(
                            text=text.rstrip(":").strip(),
                            anchor=str(getattr(cell, "coordinate", "")),
                        )
                    )
            else:
                lines: list[str] = []
                for r in range(run_start, run_end + 1):
                    parts = [
                        rows_data[r][c]
                        for c in range(start_col, end_col + 1)
                        if rows_data[r][c]
                    ]
                    line = " | ".join(parts)
                    if line:
                        lines.append(line)
                if lines:
                    blocks.append(
                        ParagraphBlock(
                            text="\n\n".join(lines),
                            anchor=f"{get_column_letter(start_col)}{run_start}",
                        )
                    )

        # Trim col counts in the built tables
        for block in blocks:
            if isinstance(block, TableBlock):
                block.headers = block.headers[:effective_max_col]
                block.rows = [row[:effective_max_col] for row in block.rows]

        return NarrativeSheet(
            name=sheet_name,
            index=index,
            blocks=blocks,
            images=images,
            raw_dimensions=raw_dimensions,
        )

    def _load_row_data(
        self,
        worksheet: object,
        max_row: int,
        max_col: int,
    ) -> list[list[str]]:
        """Read all cells once into a 2D array of strings."""
        data: list[list[str]] = [["" for _ in range(max_col + 1)] for _ in range(max_row + 1)]
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                data[r][c] = _coerce_to_string(worksheet.cell(row=r, column=c))
        return data

    def _collect_row_runs(
        self,
        rows_data: list[list[str]],
    ) -> list[tuple[int, int, int, int]]:
        """Group consecutive populated rows into rectangular runs."""
        max_row = len(rows_data) - 1
        max_col = len(rows_data[0]) - 1
        runs: list[tuple[int, int, int, int]] = []
        run_start = 0
        run_start_col = 0
        run_end_col = 0

        def _flush(end_row: int) -> None:
            nonlocal run_start
            if run_start and end_row >= run_start:
                runs.append((run_start, end_row, run_start_col, run_end_col))
            run_start = 0

        for row_idx in range(1, max_row + 1):
            non_empty_cols = [
                c for c in range(1, max_col + 1) if rows_data[row_idx][c]
            ]
            if not non_empty_cols:
                _flush(row_idx - 1)
                continue
            min_col = min(non_empty_cols)
            max_used_col = max(non_empty_cols)
            if run_start == 0:
                run_start = row_idx
                run_start_col = min_col
                run_end_col = max_used_col
                continue
            if min_col <= run_end_col and max_used_col >= run_start_col:
                run_start_col = min(run_start_col, min_col)
                run_end_col = max(run_end_col, max_used_col)
            else:
                _flush(row_idx - 1)
                run_start = row_idx
                run_start_col = min_col
                run_end_col = max_used_col

        _flush(max_row)
        return runs

    def _try_group_sparse_rows(
        self,
        rows_data: list[list[str]],
        run_start: int,
        run_end: int,
        start_col: int,
        end_col: int,
    ) -> tuple[int, int, int, int, set[int]] | None:
        """Try to group sparse rows that share a similar column footprint.

        Returns an expanded rectangle or None if grouping fails.
        """
        row_count = run_end - run_start + 1
        if row_count < 4:
            return None

        footprints: list[set[int]] = []
        for r in range(run_start, run_end + 1):
            cols = {c for c in range(start_col, end_col + 1) if rows_data[r][c]}
            if cols:
                footprints.append((r, cols))

        if len(footprints) < 3:
            return None

        # Common columns across most rows
        all_cols = footprints[0][1]
        for _, cols in footprints[1:]:
            all_cols = all_cols & cols
        if len(all_cols) < 2:
            return None

        return (run_start, run_end, start_col, end_col, all_cols)

    def _density_for_run(
        self,
        rows_data: list[list[str]],
        run_start: int,
        run_end: int,
        start_col: int,
        end_col: int,
    ) -> float:
        cells = 0
        populated = 0
        for r in range(run_start, run_end + 1):
            for c in range(start_col, end_col + 1):
                cells += 1
                if rows_data[r][c]:
                    populated += 1
        return populated / cells if cells else 0.0

    def _build_table_block_from_data(
        self,
        rows_data: list[list[str]],
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        worksheet: object,
    ) -> TableBlock:
        col_limit = self._config.max_cols or self._config.default_table_max_cols
        end_col = min(end_col, col_limit)

        rows_text: list[list[str]] = []
        anchor_start = ""
        anchor_end = ""
        for row_idx in range(start_row, end_row + 1):
            row_values = [rows_data[row_idx][c] for c in range(start_col, end_col + 1)]
            if any(v for v in row_values):
                anchor_start = anchor_start or f"{get_column_letter(start_col)}{row_idx}"
                anchor_end = f"{get_column_letter(end_col)}{row_idx}"
                rows_text.append(row_values)

        headers: list[str] = []
        body_rows: list[list[str]] = []
        if rows_text:
            first = rows_text[0]
            rest = rows_text[1:]
            if self._looks_like_header(worksheet, start_row, start_col, end_col):
                headers = first
                body_rows = rest
            else:
                body_rows = rows_text

        if self._config.max_rows is not None:
            body_rows = body_rows[: self._config.max_rows]

        return TableBlock(
            headers=headers,
            rows=body_rows,
            anchor_start=anchor_start,
            anchor_end=anchor_end,
        )

    def _detect_single_table(
        self,
        worksheet: object,
        max_row: int,
        effective_max_col: int,
        populated_rows: int,
    ) -> tuple[int, int, int, int] | None:
        """Return (start_row, end_row, start_col, end_col) if the sheet is one big table."""
        best_run = 0
        best_bounds: tuple[int, int, int, int] | None = None
        current_run_start = 0
        current_density_window: list[float] = []

        for row_idx in range(1, max_row + 1):
            non_empty = sum(
                1
                for col_idx in range(1, effective_max_col + 1)
                if getattr(worksheet.cell(row=row_idx, column=col_idx), "value", None)
                not in (None, "")
            )
            if non_empty == 0:
                current_run_start = 0
                current_density_window = []
                continue

            density = non_empty / effective_max_col
            if current_run_start == 0:
                current_run_start = row_idx
                current_density_window = [density]
            else:
                current_density_window.append(density)
                if len(current_density_window) > 5:
                    current_density_window.pop(0)

            run_length = row_idx - current_run_start + 1
            avg_density = sum(current_density_window) / len(current_density_window)
            if (
                run_length >= TABLE_REGION_MIN_ROWS
                and avg_density >= TABLE_REGION_MIN_DENSITY
                and run_length > best_run
            ):
                best_run = run_length
                best_bounds = (
                    current_run_start,
                    row_idx,
                    1,
                    effective_max_col,
                )

        if best_bounds is None:
            return None
        if best_run / max(populated_rows, 1) < TABLE_DOMINANCE_THRESHOLD:
            return None
        return best_bounds

    def _build_table_block(
        self,
        worksheet: object,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
    ) -> TableBlock:
        """Build a :class:`TableBlock` from a rectangular region."""
        col_limit = self._config.max_cols or self._config.default_table_max_cols
        end_col = min(end_col, col_limit)

        rows_text: list[list[str]] = []
        anchor_start = ""
        anchor_end = ""
        for row_idx in range(start_row, end_row + 1):
            row_values: list[str] = []
            for col_idx in range(start_col, end_col + 1):
                row_values.append(_coerce_to_string(worksheet.cell(row=row_idx, column=col_idx)))
            if any(v for v in row_values):
                anchor_start = anchor_start or f"{get_column_letter(start_col)}{row_idx}"
                anchor_end = f"{get_column_letter(end_col)}{row_idx}"
                rows_text.append(row_values)

        headers: list[str] = []
        body_rows: list[list[str]] = []
        if rows_text:
            first = rows_text[0]
            rest = rows_text[1:]
            if self._looks_like_header(worksheet, start_row, start_col, end_col):
                headers = first
                body_rows = rest
            else:
                body_rows = rows_text

        if self._config.max_rows is not None:
            body_rows = body_rows[: self._config.max_rows]

        return TableBlock(
            headers=headers,
            rows=body_rows,
            anchor_start=anchor_start,
            anchor_end=anchor_end,
        )

    def _looks_like_header(
        self,
        worksheet: object,
        row_idx: int,
        start_col: int,
        end_col: int,
    ) -> bool:
        cell = worksheet.cell(row=row_idx, column=start_col)
        if getattr(getattr(cell, "font", None), "bold", False):
            return True
        row_texts = [
            _coerce_to_string(worksheet.cell(row=row_idx, column=col_idx))
            for col_idx in range(start_col, end_col + 1)
        ]
        non_empty = [t for t in row_texts if t]
        if not non_empty:
            return False
        # All-numeric rows are data, not headers.
        if all(self._is_numeric(t) for t in non_empty):
            return False
        avg_len = sum(len(t) for t in non_empty) / len(non_empty)
        if avg_len > 80:
            return False
        upper_ratio = sum(1 for t in non_empty if t.isupper()) / len(non_empty)
        if upper_ratio >= 0.4:
            return True
        # Camel/short tokens like Col1, Col2 → header
        if all(re.match(r"^[A-Za-z]{1,5}\d{0,3}$", t) for t in non_empty):
            return True
        # Spanish keyword headers
        if any(t.lower() in HEADER_KEYWORDS for t in non_empty):
            return True
        return False

    @staticmethod
    def _is_numeric(text: str) -> bool:
        stripped = text.replace(".", "").replace(",", "").replace("-", "").replace("%", "")
        return stripped.isdigit()

    def _count_populated_rows(
        self,
        worksheet: object,
        max_row: int,
        max_col: int,
    ) -> int:
        populated = 0
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                if getattr(worksheet.cell(row=row_idx, column=col_idx), "value", None) not in (
                    None,
                    "",
                ):
                    populated += 1
                    break
        return populated

    def _density(
        self,
        worksheet: object,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
    ) -> float:
        cells = 0
        populated = 0
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cells += 1
                if getattr(worksheet.cell(row=row_idx, column=col_idx), "value", None) not in (
                    None,
                    "",
                ):
                    populated += 1
        return populated / cells if cells else 0.0

    def _extract_images(self, worksheet: object) -> list[ImageBlock]:
        if not self._config.extract_images or self._asset_exporter is None:
            return []

        images: list[ImageBlock] = []
        for index, image in enumerate(getattr(worksheet, "_images", []), start=1):
            try:
                data = image._data()  # noqa: SLF001 - openpyxl internal API
            except Exception:  # noqa: BLE001
                continue
            filename = self._asset_exporter.export(f"img_{index:03d}.png", data)
            anchor_cell = self._image_anchor(image)
            alt_text = str(getattr(image, "alt", "") or "")
            images.append(
                ImageBlock(
                    filename=filename,
                    alt_text=alt_text,
                    anchor_cell=anchor_cell,
                )
            )
        return images

    @staticmethod
    def _image_anchor(image: object) -> str:
        anchor = getattr(image, "anchor", None)
        if anchor is None:
            return ""
        anchor_from = getattr(anchor, "_from", None)
        if anchor_from is None:
            return ""
        col = int(getattr(anchor_from, "col", 0)) + 1
        row = int(getattr(anchor_from, "row", 0)) + 1
        return f"{get_column_letter(col)}{row}"


__all__ = ["XlsxParser"]
