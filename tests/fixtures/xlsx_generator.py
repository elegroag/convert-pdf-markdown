"""Helper to build XLSX files for tests."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font


def _minimal_png() -> bytes:
    return (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
        b"\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx"
        b"\x9cc\xf8\x0f\x00\x00\x01\x01\x00\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82"
    )


def build_simple_xlsx(
    path: Path | None = None,
    sheets: tuple[str, ...] = ("Resumen",),
    rows_per_sheet: int = 5,
    cols: int = 4,
) -> Path | BytesIO:
    """Build a workbook with tabular data."""
    workbook = Workbook()
    first = True
    for sheet_name in sheets:
        if first:
            worksheet = workbook.active
            worksheet.title = sheet_name
            first = False
        else:
            worksheet = workbook.create_sheet(sheet_name)

        for col in range(1, cols + 1):
            worksheet.cell(row=1, column=col, value=f"Col{col}")
        for row in range(2, rows_per_sheet + 1):
            for col in range(1, cols + 1):
                worksheet.cell(row=row, column=col, value=f"R{row}C{col}")

    if path is not None:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(path))
        return path

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_xlsx_with_images(path: Path | None = None) -> Path:
    """Build a workbook with a minimal embedded PNG image."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Imagenes"
    worksheet["A1"] = "Con imagen"
    png_bytes = _minimal_png()
    image = XLImage(BytesIO(png_bytes))
    worksheet.add_image(image, "B2")

    target = path or Path("with_image.xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(target))
    return target


def build_xlsx_with_formulas(path: Path | None = None) -> Path:
    """Build a workbook with spreadsheet formulas."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Formulas"
    for row in range(1, 6):
        worksheet.cell(row=row, column=1, value=row)
    worksheet["B1"] = "Total"
    worksheet["B2"] = "=SUM(A1:A5)"
    worksheet["B3"] = "=A1*A2"

    target = path or Path("with_formulas.xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(target))
    return target


def build_xlsx_with_merges(path: Path | None = None) -> Path:
    """Build a workbook with merged header cells."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Merges"
    worksheet.merge_cells("A1:C1")
    cell = worksheet["A1"]
    cell.value = "Encabezado combinado"
    cell.font = Font(bold=True)
    worksheet["A2"] = "A"
    worksheet["B2"] = "B"
    worksheet["C2"] = "C"

    target = path or Path("with_merges.xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(target))
    return target


def build_xlsx_multi_sheet(path: Path | None = None) -> Path:
    """Build a workbook with multiple sheets, including one empty sheet."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Resumen"
    for col in range(1, 4):
        worksheet.cell(row=1, column=col, value=f"Col{col}")
    for row in range(2, 4):
        for col in range(1, 4):
            worksheet.cell(row=row, column=col, value=f"R{row}C{col}")

    datos = workbook.create_sheet("Datos")
    for col in range(1, 4):
        datos.cell(row=1, column=col, value=f"Col{col}")
    for row in range(2, 4):
        for col in range(1, 4):
            datos.cell(row=row, column=col, value=f"R{row}C{col}")

    workbook.create_sheet("Vacia")

    target = path or Path("multi.xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(target))
    return target


def build_xlsx_with_hyperlink(path: Path | None = None) -> Path:
    """Build a workbook with a hyperlink cell."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Links"
    worksheet["A1"] = "Sitio"
    worksheet["A2"].hyperlink = "https://example.com"
    worksheet["A2"].value = "Example"
    worksheet["A2"].style = "Hyperlink"

    target = path or Path("with_hyperlink.xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(target))
    return target


def build_xlsx_with_narrative(path: Path | None = None) -> Path:
    """Build an "informe" with key/value pairs followed by a table."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Informe"

    bold = Font(bold=True)
    worksheet["C2"] = "OBJETIVO:"
    worksheet["C2"].font = bold
    worksheet["D2"] = "Evaluar el SGSI de la organización."
    worksheet["C4"] = "Jefe del departamento:"
    worksheet["C4"].font = bold
    worksheet["D4"] = "Ricardo Liévano"

    # Table region with header
    for col_idx, header in enumerate(["Item", "Recomendación"], start=1):
        cell = worksheet.cell(row=14, column=col_idx, value=header)
        cell.font = bold
    worksheet.cell(row=15, column=1, value=1)
    worksheet.cell(row=15, column=2, value="Cumplir el manual")
    worksheet.cell(row=16, column=1, value=2)
    worksheet.cell(row=16, column=2, value="Actualizar matriz")

    target = path or Path("informe.xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(target))
    return target
